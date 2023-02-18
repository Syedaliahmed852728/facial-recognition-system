import os
import cv2
import cvzone
import face_recognition
import numpy as np
import firebase_admin
from firebase_admin import credentials
from firebase_admin import db
from firebase_admin import storage
import pickle
from datetime import datetime
# import datetime
from openpyxl import load_workbook, Workbook

cred = credentials.Certificate("serviceAccountKey.json")
firebase_admin.initialize_app(cred, {
    'databaseURL': "https://facerecognitionrealtime-40d84-default-rtdb.firebaseio.com/",
    'storageBucket': 'facerecognitionrealtime-40d84.appspot.com'
})

bucket = storage.bucket()

cap = cv2.VideoCapture(0)
cap.set(3, 640)
cap.set(4, 480)

imageBackGround = cv2.imread("resources/backgroundImage.png")

# importing the images for example active,alreadyMarked etc

folderModePath = 'resources/modes'
modePathList = os.listdir(folderModePath)
imageModeList = []

for path in modePathList:
    imageModeList.append(cv2.imread(os.path.join(folderModePath, path)))

    # print(len(imageModeList))
    # load the encoding file

    print("loading encode File....")
    file = open('EncodeFile.p', 'rb')
    encodingListKnownWithIds = pickle.load(file)
    file.close()
    encodeListKnown, studentIds = encodingListKnownWithIds
    print("encode File loaded")

    modeType = 0
    counter = 0
    id = -1
    imageStudent = []
# --------------------------------------------------------------------------------
# choose excel file for attendance
opt = 0
while opt < 1 or opt >7:
    opt = int(input("Select Subject:\n1.DSA\n2.CN\n3.OOAD\n4.SRE\n5.BPE\n6.OOP\nEnter the respective number: "))
subjects = {1: "DSA.xlsx", 2: "CN.xlsx", 3: "OOAD.xlsx", 4: "SRE.xlsx", 5: "BPE.xlsx", 6: "OOP.xlsx"}
fileName = subjects.get(opt)
# dictionary to enter the correct sheet
date = datetime.now()
monthNumber = date.month
monthDict = {1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
         7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'}
myMonth = monthDict.get(monthNumber)
key = input("press any key to continue...")
# ----------------------------------------------------------------------------------
while True:

    success, image = cap.read()
    imageSmall = cv2.resize(image, (0, 0), None, 0.25, 0.25)
    imageSmall = cv2.cvtColor(imageSmall, cv2.COLOR_BGR2RGB)
    faceCurrFram = face_recognition.face_locations(imageSmall)
    encodeCurrFrame = face_recognition.face_encodings(imageSmall, faceCurrFram)
    imageBackGround[162:162 + 480, 55:55 + 640] = image
    imageBackGround[44:44 + 633, 808:808 + 414] = imageModeList[modeType]

    if faceCurrFram:
        for faceEncode, faceLoc in zip(encodeCurrFrame, faceCurrFram):
            matches = face_recognition.compare_faces(encodeListKnown, faceEncode)
            faceDist = face_recognition.face_distance(encodeListKnown, faceEncode)
            matchIndex = np.argmin(faceDist)

            if matches[matchIndex]:
                y1, x2, y2, x1 = faceLoc
                y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                bbox = 55 + x1, 162 + y1, x2 - x1, y2 - y1
                imageBackGround = cvzone.cornerRect(imageBackGround, bbox, rt=0)
                id = studentIds[matchIndex]
                if counter == 0:
                    cvzone.putTextRect(imageBackGround, "Loading", (275, 400))
                    cv2.imshow("face_Attendance", imageBackGround)
                    cv2.waitKey(3)
                    counter = 1
                    modeType = 1

        if counter != 0:
            if counter == 1:
                # get the data
                studentInfo = db.reference(f'students/{id}').get()
                # ------------------------------------------------------
                # marking this student as present in selected excel file
                rollNumber = db.reference(f'students/{id}').key
                # print(studentInfo)
                wb = load_workbook(fileName)
                ws = wb[myMonth]
                maxRows = ws.max_row
                maxColumns = ws.max_column
                column = ws['A']
                index = 0
                today = date.day
                rowNumber = 2 + today
                print('rowNumber: ', rowNumber)
                for cell in column:
                    index += 1
                    if cell.value == int(rollNumber):
                        print('found....')
                        target = ws[index]
                        target[rowNumber].value = "Present"
                        wb.save(filename=fileName)
                        break
                # print(studentInfo['Name'], " is present")
                # ----------------------------------------------------------
                # get the image from the storage
                blob = bucket.get_blob(f'image/{id}.png')
                array = np.frombuffer(blob.download_as_string(), np.uint8)
                imageStudent = cv2.imdecode(array, cv2.COLOR_BGRA2BGR)
                # upadate data of attendace
                datetimeObject = datetime.strptime(studentInfo['Last_attendace_Time'],
                                                   "%Y-%m-%d  %H:%M:%S")
                secondElapesed = (datetime.now() - datetimeObject).total_seconds()
                if secondElapesed > 30:
                    ref = db.reference(f'students/{id}')
                    studentInfo['total_attendance'] += 1
                    ref.child('total_attendance').set(studentInfo['total_attendance'])
                    ref.child('Last_attendace_Time').set(datetime.now().strftime("%Y-%m-%d  %H:%M:%S"))
                else:
                    modeType = 3
                    counter = 0
                    imageBackGround[44:44 + 633, 808:808 + 414] = imageModeList[modeType]

            if modeType != 3:
                if 10 < counter <= 20:
                    modeType = 2
                    imageBackGround[44:44 + 633, 808:808 + 414] = imageModeList[modeType]
                if counter <= 10:
                    cv2.putText(imageBackGround, str(studentInfo['total_attendance']), (861, 125),
                                cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 1)
                    cv2.putText(imageBackGround, str(studentInfo['Major']), (1006, 550),
                                cv2.FONT_HERSHEY_COMPLEX, 0.5, (255, 255, 255), 1)
                    cv2.putText(imageBackGround, str(id), (1006, 493),
                                cv2.FONT_HERSHEY_COMPLEX, 0.5, (255, 255, 255), 1)
                    cv2.putText(imageBackGround, str(studentInfo['Standing']), (910, 625),
                                cv2.FONT_HERSHEY_COMPLEX, 0.6, (100, 100, 100), 1)
                    cv2.putText(imageBackGround, str(studentInfo['Starting_Year']), (1125, 625),
                                cv2.FONT_HERSHEY_COMPLEX, 0.6, (100, 100, 100), 1)
                    (w, h), _ = cv2.getTextSize(studentInfo['Name'], cv2.FONT_HERSHEY_COMPLEX, 1, 1)
                    offset = (414 - w) // 2
                    cv2.putText(imageBackGround, str(studentInfo['Name']), (808 + offset, 445),
                                cv2.FONT_HERSHEY_COMPLEX, 1, (50, 50, 50), 1)
                    imageBackGround[175:175 + 216, 909:909 + 216] = imageStudent

                counter += 1

                if counter >= 20:
                    counter = 0
                    modeType = 0
                    studentInfo = []
                    imageStudent = []
                    imageBackGround[44:44 + 633, 808:808 + 414] = imageModeList[modeType]

    else:
        modeType = 0
        counter = 0

    cv2.imshow("face_Attendance", imageBackGround)
    cv2.waitKey(1)
